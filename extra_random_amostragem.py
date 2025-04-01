import csv
import json
import random
import os
import subprocess

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def write_in_excel(caminho_excel, project, review, revision, clone_diff, clone_before, aba="Amostra"):

    workbook = load_workbook(caminho_excel)
    if aba not in workbook.sheetnames:
        sheet = workbook.create_sheet(aba)
    else:
        sheet = workbook[aba]


    nova_linha = [project, review, revision, clone_diff, clone_before, "", ""]

    sheet.append(nova_linha)

    workbook.save(caminho_excel)

def ler_csv(caminho_arquivo):
    try:
        with open(caminho_arquivo, mode='r', newline='', encoding='utf-8') as file:
            leitor = csv.reader(file)
            linhas = list(leitor)  
            return linhas
    except FileNotFoundError:
        return f"Arquivo '{caminho_arquivo}' não encontrado."
    except Exception as e:
        return f"Erro ao ler o arquivo: {e}"

def encontrar_arquivo_csv(caminho_pasta, parte_nome):

    conteudo_pasta = os.listdir(caminho_pasta)
    revision = parte_nome.split("_")[3]

    for arquivo in conteudo_pasta:
        if arquivo.endswith('.csv') and parte_nome in arquivo and arquivo.split("_")[3] == revision:
            return ler_csv(os.path.join(caminho_pasta, arquivo))

def ler_arquivo_java(caminho_arquivo):
    encodings = ["utf-8", "latin1", "iso-8859-1", "utf-16", "cp1252"]
    for encoding in encodings:
        try:
            with open(caminho_arquivo, "r", encoding=encoding) as arquivo:
                conteudo = arquivo.read()
            print(f"Arquivo lido com sucesso usando o encoding: {encoding}")
            return conteudo
        except UnicodeDecodeError:
            print(f"Erro ao tentar ler o arquivo com o encoding: {encoding}")
        except FileNotFoundError:
            print(f"Arquivo {caminho_arquivo} não encontrado.")
            return None
    print("Falha ao ler o arquivo com os encodings fornecidos.")
    return None


def ler_linhas_arquivo(caminho_arquivo, linha_inicio, linha_fim):
    encodings = ["utf-8", "latin1", "iso-8859-1", "utf-16", "cp1252"]

    for encoding in encodings:
        try:
            with open(caminho_arquivo, "r", encoding=encoding) as arquivo:
                linhas = arquivo.readlines()  # Lê todas as linhas do arquivo
                if linha_inicio < 1 or linha_fim > len(linhas):
                    print(f"Linhas inválidas. O arquivo tem {len(linhas)} linhas.")
                    return None
                return ''.join(linhas[linha_inicio - 1:linha_fim])  
            print(f"Arquivo lido com sucesso usando o encoding: {encoding}")
        except UnicodeDecodeError:
            print(f"Erro ao tentar ler o arquivo com o encoding: {encoding}")
        except FileNotFoundError:
            print(f"Arquivo {caminho_arquivo} não encontrado.")
            return None
    print("Falha ao ler o arquivo com os encodings fornecidos.")
    return None

# Script starts here
# Projeto ao qual quero pegar a amostra
project = "platform.ui"
# Qnt da amostra
qnt_amostra = 1
#Qnt Piloto
qnt_piloto = 0
# O csv sem especificação é igual ao equals_revisions
caminho_type_clones_equals_reviews = "type_clones/" + project + "_equals_reviews.json"
caminho_type_clones_original = "type_clones/" + project + "(original).csv"
caminho_metadata = "metadata/" + project + ".csv"
caminho_repo = "git_repos/" + project
caminho_excel = "amostra_excel/Amostras.xlsx"

if not os.path.exists("amostra_excel"):
    os.makedirs("amostra_excel")

if not os.path.exists(caminho_excel):
    # Criar um novo arquivo Excel
    workbook = Workbook()

    # Criar a aba "Amostra"
    sheet_amostra = workbook.active
    sheet_amostra.title = "Amostra"

    # Adicionar cabeçalho para "Amostra"
    header_amostra = ["project", "review", "revision", "clone_diff", "clone_before", "is clone?", "type clone"]
    sheet_amostra.append(header_amostra)

    # Configurar validação de dados para "is clone?" e "type clone" na aba "Amostra"
    dv_is_clone = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_type_clone = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    sheet_amostra.add_data_validation(dv_is_clone)
    sheet_amostra.add_data_validation(dv_type_clone)
    dv_is_clone.add("F2:F1048576")  # Coluna "is clone?" (começa na linha 2)
    dv_type_clone.add("G2:G1048576")  # Coluna "type clone" (começa na linha 2)

    # Criar a aba "Piloto"
    sheet_piloto = workbook.create_sheet(title="Piloto")

    # Adicionar cabeçalho para "Piloto"
    header_piloto = ["project", "review", "revision", "clone_diff", "clone_before", "is clone?", "type clone"]
    sheet_piloto.append(header_piloto)

    # Configurar validação de dados para "is clone?" e "type clone" na aba "Piloto"
    sheet_piloto.add_data_validation(dv_is_clone)
    sheet_piloto.add_data_validation(dv_type_clone)
    dv_is_clone.add("F2:F1048576")  # Coluna "is clone?" (começa na linha 2)
    dv_type_clone.add("G2:G1048576")  # Coluna "type clone" (começa na linha 2)

    # Salvar o arquivo
    workbook.save(caminho_excel)

with open(caminho_type_clones_original, newline='', encoding='utf-8') as csvfile:
    leitor_csv = csv.reader(csvfile)
    contador_amostra = sum(1 for _ in leitor_csv)

with open(caminho_type_clones_equals_reviews, 'r', encoding='utf-8') as arquivo:
    linhas_dict = json.load(arquivo)

# Blocos sorteados para a amostra
blocos_sorteados = []
numeros_sorteados_amostra = []
# Blocos sorteados para o piloto
blocos_piloto = []
numeros_sorteados_piloto = []

# Sorteando os blocos para a amostra e piloto
while len(numeros_sorteados_amostra) + len(numeros_sorteados_piloto) != (qnt_amostra + qnt_piloto):
    num_random = random.randint(2, contador_amostra + 1)
    if num_random not in numeros_sorteados_amostra + numeros_sorteados_piloto:
        if str(num_random) in linhas_dict:
            bloco = linhas_dict[str(num_random)][2]
            clone_false = linhas_dict[str(num_random)][3]
            clone_from_test = linhas_dict[str(num_random)][4]
            if bloco not in blocos_sorteados and clone_false == "0" and clone_from_test == "0":
                if len(numeros_sorteados_amostra) < qnt_amostra:
                    blocos_sorteados.append(bloco)
                    numeros_sorteados_amostra.append(num_random)
                else:
                    blocos_piloto.append(bloco)
                    numeros_sorteados_piloto.append(num_random)

# Carregar as linhas do CSV original e selecionar os blocos sorteados
with open(caminho_type_clones_original, newline='') as csvfile:
    leitor_csv = csv.reader(csvfile, delimiter='\t')
    header = next(leitor_csv)
    linhas_dict_original = {}
    for idx, linha in enumerate(leitor_csv, start=2):  # Enumerar começa de 1
        if idx in numeros_sorteados_amostra + numeros_sorteados_piloto:
            linhas_dict_original[idx] = linha  # Associa o índice à linha

lista_blocos_sorteados = []
for linha_dict_original in linhas_dict_original:
    linha_splitted = linhas_dict_original[linha_dict_original][0].split(",")
    lista_blocos_sorteados.append(linha_splitted)

with open(caminho_metadata, newline='') as csvfile:
    leitor_csv = csv.reader(csvfile, delimiter='\t')
    header = next(leitor_csv)
    before_ids = []
    for linha in leitor_csv:
        linha = linha[0].split(",")
        for bloco in lista_blocos_sorteados:
            if linha[1] == bloco[0] and linha[2] == bloco[1]:
                L = list(linha)
                L.append(bloco[2])
                if L not in before_ids:
                    before_ids.append(L)

count = 1
for id in before_ids:
    review = id[1]
    revision = id[2]
    block = id[11]
    number_block = block.split(".")[0].split("_")[1]
    caminho_bloco_diff = "blocks/" + project + "/" + review + "_" + revision + "/" + block
    conteudo_bloco_diff = ler_arquivo_java(caminho_bloco_diff)
    before_id = id[9]
    caminho_search_results = "search_results/"
    arquivo_search_results = project + "_before_" + review + "_" + revision
    conteudo_completo_search_results = encontrar_arquivo_csv(caminho_search_results, arquivo_search_results)
    split_caminho_before = "before_" + review + "_" + revision + "/"
    for linha_csv in conteudo_completo_search_results:
        if number_block == linha_csv[0].split(".")[0].split("block")[1].strip():
            print(review, revision, number_block, before_id)
            caminho_clone_before = "git_repos/" + project + "/" + linha_csv[1].split(split_caminho_before)[1].rsplit("_", 1)[0]
            linha_metodo_inicio = linha_csv[1].split(split_caminho_before)[1].rsplit("_", 1)[1].split("#")[1]
            linha_metodo_final = linha_csv[1].split(split_caminho_before)[1].rsplit("_", 1)[1].split("#")[2]
            # Comando para resetar o repositório para um commit específico
            reset_command = f"git -C {caminho_repo} reset --hard {before_id}"
            print("Amostra número: " + str(count))
            # Executa o comando git reset usando subprocess
            subprocess.run(reset_command, shell=True, check=True)
            conteudo_bloco_before = ler_linhas_arquivo(caminho_clone_before, int(linha_metodo_inicio), int(linha_metodo_final))
            if count <= qnt_amostra:
                write_in_excel(caminho_excel, project, review, revision, conteudo_bloco_diff, conteudo_bloco_before, aba="Amostra")
            else:
                write_in_excel(caminho_excel, project, review, revision, conteudo_bloco_diff, conteudo_bloco_before, aba="Piloto")
            break
    count += 1
